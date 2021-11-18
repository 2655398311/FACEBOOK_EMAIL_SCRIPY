import os
import time
import random

from selenium import webdriver
import json
import requests
import re
from openpyxl import load_workbook, Workbook
# 读取昵称
def read_name(file_name):
    # 获取昵称
    wb = load_workbook(file_name)
    ws = wb.active
    data = []
    for row in ws:
        row_value = []
        for cell in row:
            value = cell.value
            row_value.append(value)
        data.append(row_value)
    return data  # 返回用户昵称的列表

def get_page(name, cookies):
    base_url = 'https://www.facebook.com/search/top/?q='
    # 请求头
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cache-control': 'max-age=0',
        'cookie': cookies,
        'referer': 'https://www.facebook.com/',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'
    }
    proxies = {"http": f"http://127.0.0.1:10809", "https": f"http://127.0.0.1:10809"}
    url = base_url + name.replace(' ', '%20')
    print(url)

    res = requests.get(url, headers=headers,timeout = 160,proxies =proxies)

    res.encoding = 'utf-8'  # 设置编码格式
    res_text = res.text.replace(r'\/', '/')
    try:
        text = re.findall(r'"search_cta_model":{"place":(.*?)}', res_text)[0]

        res = re.findall(r'"url":"https://www\.facebook\.com/(.*?)"', text)
        print("2*2" * 30, res)

        for r in res:
            if r:
                page_url = r'https://www.facebook.com/' + r  # 拼接用户主页的URL

                break
        else:
            page_url = ''
    except:
        # print('error:', url)
        res = re.findall(r'"url":"https://www\.facebook\.com/(.*?)/"', res_text)
        print("1*1" * 30, res)

        for r in res:
            if r:
                page_url = r'https://www.facebook.com/' + r  # 拼接用户主页的URL

                break
        else:
            page_url = ''
    # print(name, url, page_url)
    return page_url

def get_mail(page_url, cookies):
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cache-control': 'max-age=0',
        'cookie': cookies,
        'referer': page_url,
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'
    }
    while True:
        try:
            proxies = {"http": f"http://127.0.0.1:10809", "https": f"http://127.0.0.1:10809"}

            res = requests.get(page_url, headers=headers,proxies = proxies)
            break
        except Exception as e:
            print('error-2:', e)
    res.encoding = 'utf-8'
    text = res.text

    # 提取需要的数据'
    res = re.findall('"email":{"text":"(.*?)"', text)
    for r in res:
        if r:
            if r"\u0040" in r:
                email_address = r.replace(r"\u0040", "@")  # 邮箱地址
                break
    else:
        email_address = ''
    return email_address

# 获取已经爬取的昵称
def names():
    names = []
    if os.path.exists("已获取.txt"):
        with open('已获取.txt', 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                name = line.strip()
                names.append(name)
    return names

def aa():
    driver = webdriver.Chrome()
    driver.get('https://www.facebook.com/')

    time.sleep(2)
    user = driver.find_element_by_id("email")
    user.send_keys('ebaguc@mail.com.tr')
    # 找到密码输入密码
    driver.find_element_by_id("pass").send_keys('3lhEo5cAx')
    # 点击登录按钮实现登录
    driver.find_element_by_xpath("//button").click()

    time.sleep(3)

    # aa = requests.get("http://198.13.54.13/google/getCode.php?secret=MPNM3BW3R4HXDZCRBCZXZZNYMN7WR4UM")
    aa = requests.get("http://198.13.54.13/google/getCode.php?secret=CTUJC4MWMFPAJI45FS4VCKTSQTPMPQTO")
    code = aa.text

    driver.find_element_by_id("approvals_code").send_keys(code)
    driver.find_element_by_xpath("//button").click()
    time.sleep(2)
    driver.find_element_by_xpath("//button").click()


    Cookie = driver.get_cookies()
    strr = ''
    for c in Cookie:
        strr += c['name']
        strr += '='
        strr += c['value']
        strr += ';'
    full_cookie = strr
    # time.sleep(899999)
    return full_cookie

def save_mail(file_name, item):
    # 判断文件是否存在
    result_file = file_name.replace('.xlsx', '_result_usa_1029.xlsx')
    # 如果文件不存在就新建一个工作簿
    if os.path.exists(result_file):
        wb = load_workbook(result_file)
    else:
        wb = Workbook()
    ws = wb.active  # 获取工作表
    ws.append(item)  # 添加数据
    wb.save(result_file)  # 保存文件
    wb.close()  # 关闭文件

if __name__ == '__main__':
    names = names()
    file_name = r'USA(1)(1).xlsx'
    data = read_name(file_name)  # 读取用户昵称
    # name = "PNEUS METRO INC."
    cookies = aa()
    # print(cookies)
    # for row in data:
    #     name = row[2].strip()
    #     if name in names:
    #         continue
    #     cc = get_page(name,cookies)
    #     email = get_mail(cc, cookies)  # 获取邮箱
    #     row.append(email)
    #     row.append(cc)
    #     save_mail(file_name, row)
    #     with open("已获取.txt", 'a', encoding='utf-8') as f:
    #         f.write(name + '\n')
    #     print(name, "url:", cc, 'email:', email)
    #     # sleeptime = random.randint(1, 30)
    #     time.sleep(60)
